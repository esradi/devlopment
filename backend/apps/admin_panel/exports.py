import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class AdminExporter:
    """
    Utility class to generate styled Excel files from querysets
    """

    @staticmethod
    def export_users_to_excel(queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Students Export"

        # Headers
        headers = [
            "ID", "First Name", "Last Name", "Email", 
            "Domain", "Speciality", "Status", "Date Joined"
        ]
        
        # Style Headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="9E59FF", end_color="9E59FF", fill_type="solid") # Elegant Purple
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data
        for row, user in enumerate(queryset, 2):
            try:
                domain = user.student_profile.domain or ''
                speciality = user.student_profile.speciality or ''
            except Exception:
                domain = ''
                speciality = ''

            ws.cell(row=row, column=1, value=user.id)
            ws.cell(row=row, column=2, value=user.first_name)
            ws.cell(row=row, column=3, value=user.last_name)
            ws.cell(row=row, column=4, value=user.email)
            ws.cell(row=row, column=5, value=domain)
            ws.cell(row=row, column=6, value=speciality)
            ws.cell(row=row, column=7, value="Active" if getattr(user, 'is_active', False) else "Inactive")
            ws.cell(row=row, column=8, value=user.date_joined.strftime("%Y-%m-%d"))

        # Column width
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def export_applications_to_excel(queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Applications Monitoring"

        # Headers
        headers = [
            "ID", "Candidate Name", "Candidate Email", "Offer Title", 
            "Company", "Status", "Applied Date", "Last Updated"
        ]
        
        # Style Headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="9E59FF", end_color="9E59FF", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data
        for row, app in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=app.id)
            ws.cell(row=row, column=2, value=app.student.user.get_full_name())
            ws.cell(row=row, column=3, value=app.student.user.email)
            ws.cell(row=row, column=4, value=app.offer.title)
            ws.cell(row=row, column=5, value=app.company.company_name)
            ws.cell(row=row, column=6, value=app.status.upper())
            ws.cell(row=row, column=7, value=app.created_at.strftime("%Y-%m-%d %H:%M"))
            ws.cell(row=row, column=8, value=app.updated_at.strftime("%Y-%m-%d %H:%M"))

        # Column width
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 25

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
